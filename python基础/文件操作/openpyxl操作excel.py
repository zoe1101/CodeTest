import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

'''
样式设置参考：
https://mp.weixin.qq.com/s?__biz=MzAxMjUyNDQ5OA==&mid=2653577413&idx=1&sn=8129b8339233686054f9f9a72230e4ae&chksm=806e4e78b719c76ebe73b98f9c7891bf917d2a83b428d92d4f973987d34194adb2c228792dca&scene=27
'''


class ExcelOperation():
    def __init__(self, file_path, sheet_name="Sheet1"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(self.file_path)
        if self.sheet_name is None:
            self.worksheet = self.workbook.active  # active默认读取第一个表单,使用active属性获取当前工作薄的活动表
        else:
            self.worksheet = self.workbook[self.sheet_name]  # 读取指定表单

    def change_workbook(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(self.file_path)
        if self.sheet_name not in self.workbook.sheetnames:
            self.add_sheet(self.sheet_name)
        else:
            self.worksheet = self.workbook[self.sheet_name]

    def get_rows(self):
        '''行数统计'''
        return self.worksheet.max_row

    def get_cols(self):
        '''列数统计 '''
        return self.worksheet.max_column

    def delete_by_rows_cols(self, row_list=[], cols_list=[]):
        '''
        基于ws删除一些行和一些列，注意没有备份
        :param row_list:
        :param cols_list:
        :return:
        '''
        row_list.sort(reverse=True)  # 确保大的行数首先删除
        cols_list.sort(reverse=True)  # 确保大的列数首先删除
        for row in row_list:
            self.worksheet.delete_rows(row)
        for col in cols_list:
            self.worksheet.delete_cols(col)
        self.save_excel()

    def read_row(self, row_index):
        res = []
        for col in range(self.get_cols()):
            res.append(self.read_cell(row_index, col + 1))
        return res

    def write_row(self, value, row_index):
        for col in range(len(value)):
            self.worksheet.cell(row_index, col + 1).value = value[col]
        self.save_excel()

    def read_col(self, col_index):
        res = []
        for row in range(self.get_rows()):
            res.append(self.read_cell(row + 1, col_index))
        return res

    def write_col(self, value, col_index):
        for row in range(len(value)):
            self.worksheet.cell(row + 1, col_index).value = value[row]
        self.save_excel()

    def read_cell(self, row_index, col_index):
        return self.worksheet.cell(row_index, col_index).value

    def write_cell(self, value, row_index, col_index):
        self.worksheet.cell(row_index, col_index).value = value
        self.save_excel()

    def read_area(self, row_start_index=1, row_end_index=None, col_start_index=1, col_end_index=None):
        row_end_index = self.get_rows() if not row_end_index else row_end_index
        col_end_index = self.get_cols() if not col_end_index else col_end_index

        res = []
        for row in range(row_start_index, row_end_index + 1):
            temp = []
            for col in range(col_start_index, col_end_index + 1):
                temp.append(self.read_cell(row, col))
            res.append(temp)
        return res

    def write_area(self, area_value, row_start_index=1, col_start_index=1):
        m, n = len(area_value), len(area_value[0])
        for i in range(m):
            for j in range(n):
                self.worksheet.cell(row_start_index + i, col_start_index + j).value = area_value[i][j]
        self.save_excel()

    def save_excel(self):
        """
        保存表格
        :param path: 保存的路径
        """
        self.workbook.save(self.file_path)

    def set_font(self, row_start_index, row_end_index, col_start_index, col_end_index, **kwargs):
        '''
        表格字体设置
        :return:
        '''
        for row in range(row_start_index, row_end_index + 1):
            for col in range(col_start_index, col_end_index + 1):
                self.worksheet.cell(row, col).font = Font(**kwargs)
        self.save_excel()

    def set_alignment(self, row_start_index, row_end_index, col_start_index, col_end_index, **kwargs):
        '''
        表格对齐设置
        :return:
        '''
        for row in range(row_start_index, row_end_index + 1):
            for col in range(col_start_index, col_end_index + 1):
                self.worksheet.cell(row, col).alignment = Alignment(**kwargs)
        self.save_excel()

    def set_border(self, row_start_index, row_end_index, col_start_index, col_end_index, **kwargs):
        '''
        表格边框设置
        :return:
        '''
        for row in range(row_start_index, row_end_index + 1):
            for col in range(col_start_index, col_end_index + 1):
                self.worksheet.cell(row, col).border = Border(**kwargs)
        self.save_excel()

    def set_background_colors(self, row_start_index, row_end_index, col_start_index, col_end_index, **kwargs):
        '''
        表格背景颜色设置
        :return:
        '''
        for row in range(row_start_index, row_end_index + 1):
            for col in range(col_start_index, col_end_index + 1):
                self.worksheet.cell(row, col).fill = PatternFill(**kwargs)
        self.save_excel()

    def add_sheet(self, sheet_name, idx=None):
        '''
        添加表格sheet页
        '''
        self.workbook.create_sheet(sheet_name, idx)
        self.save_excel()

    def delete_sheet(self, sheet_name):
        '''
        删除sheet页
        '''
        if sheet_name in self.workbook.sheetnames:
            print('dddddd')
            del self.workbook[sheet_name]
            self.save_excel()

    def update_sheet_name(self, ori_name, new_name):
        '''
        修改sheet名
        '''
        self.workbook[ori_name].title = new_name
        self.save_excel()


if __name__ == '__main__':
    sheet = ExcelOperation("../../data/test.xlsx", "Sheet1")
    # data = sheet.read_area(1, 1, 1, 3) #读
    # print(data)
    # sheet.write_area(data, 4, 1) #写
    # sheet.add_sheet('newsheet')
    # sheet.update_sheet_name('newsheet','newsheetu')
    # sheet.delete_sheet('newsheetu')

    # sheet.set_font(1,1,1,3,name="Arial", size=14, color="00FF0000")
    # sheet.set_alignment(1,1,1,3,horizontal='left',vertical='center')

    thin = Side(border_style="thin", color='00FF00FF')
    double = Side(border_style="double", color='00008000')
    sheet.set_border(1, 1, 1, 3, top=double, left=thin, right=thin, bottom=double)
    sheet.set_background_colors(1, 1, 1, 3, start_color='00FFFF00', end_color='00FFFF00', fill_type="solid")
